import types

import openpyxl


def separator(string, prefix='/', slicer=", "):
    """
    Function for making a dictionary that would be comfortable to use with catalog_controller

    :param string: the input line that would be separated on name, price and count
    :param prefix: bot prefix that deletes command characters from line (default '/')
    :param slicer: char set that would be sliced on parts (default ", ")
    """
    if prefix != None:
        string = string.replace(prefix + ' ', '', 1)
    string = string.split(slicer)
    data = {
        'maker': string[0].upper(),
        'taste': string[1],
        'puffs': 0,
        'price': 0,
        'count': 1,

    }
    try:
        if data['puffs'] < int(string[2]): data['puffs'] = int(string[2])
    except Exception:
        print(f"count по умолчанию {data['count']}")

    try:
        if data['price'] < int(string[3]): data['price'] = int(string[3])
    except Exception:
        print(f"price по умолчанию {data['price']}")

    try:
        if data['count'] < int(string[4]):
            data['count'] = int(string[4])
    except Exception:
        print(f"count по умолчанию {data['count']}")

    print(data)
    return data


def catalog_add(maker, taste, puffs, price=0, count=1, path="databases/catalog.xlsx"):
    """
    Function for adding new product or increase count of existing product
    :param maker: name of comp
    :param taste: taste of product
    :param puffs: count of puffs
    :param price: cost of product
    :param count: count of product on store
    :param path: path to xlsx
    :return: none
    """

    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    is_found = 0
    for i in range(1, sheet.max_row + 1):
        if maker in str(sheet.cell(i, 1).value) and taste in str(sheet.cell(i, 2).value):
            sheet.cell(i, 5).value = int(sheet.cell(i, 5).value) + count
            is_found = 1
            break

    if is_found == 0:
        placeholder = sheet.max_row + 1
        sheet.cell(placeholder, 1).value = maker
        sheet.cell(placeholder, 2).value = taste
        sheet.cell(placeholder, 3).value = puffs
        sheet.cell(placeholder, 4).value = price
        sheet.cell(placeholder, 5).value = count

    wb.save(path)
    wb.close()


def catalog_del(maker, taste, path="databases/catalog.xlsx"):
    """
    Deletes row from catalog
    :param maker:
    :param taste:
    :param path:
    :return: maker and taste string
    """
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    info = 0
    for i in range(1, sheet.max_row + 1):
        if str(maker).lower() in str(sheet.cell(i, 1).value).lower() and str(taste).lower() in str(
                sheet.cell(i, 2).value).lower():
            workcell = i
            info = str(sheet.cell(workcell, 1).value)
            sheet.cell(workcell, 1).value = ''
            sheet.cell(workcell, 2).value = ''
            sheet.cell(workcell, 3).value = ''
            sheet.cell(workcell, 4).value = ''
            sheet.cell(workcell, 5).value = ''
            break

    wb.save(path)
    wb.close()
    return info


def catalog_sell(maker, taste, count=1, path="databases/catalog.xlsx"):
    """
    Decreases count of product, if last product - deletes row
    :param maker:
    :param taste:
    :param count:
    :param path:
    :return: maker and taste string
    """
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    found = 0
    print(taste)
    # уменьшение количетва, если последний товар - удаление
    for i in range(1, sheet.max_row + 1):
        if str(maker).lower() in str(sheet.cell(i, 1).value).lower() and str(taste).lower() in str(
                sheet.cell(i, 2).value).lower():
            workcell = i
            found = f"{str(sheet.cell(workcell, 1).value)} {str(sheet.cell(workcell, 2).value)}"

            sheet.cell(workcell, 5).value = int(sheet.cell(workcell, 5).value) - count
            break

    wb.save(path)
    wb.close()
    return found


def remove_last_pos(path="databases/catalog.xlsx"):
    """
    Function removes last product row from the catalog
    :param path: path to excel table (default "catalog.excel")
    :return:
    """

    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    workcell = sheet.max_row

    sheet.cell(workcell, 1).value = ''
    sheet.cell(workcell, 2).value = ''
    sheet.cell(workcell, 3).value = ''
    sheet.cell(workcell, 4).value = ''
    sheet.cell(workcell, 5).value = ''

    wb.save(path)
    wb.close()


def get_makers(path="databases/catalog.xlsx", column=1):
    """
    Get all unic makers from xlxs (default - 1 column of each row)
    :param path:
    :return: list of makers
    """
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    workcell = sheet.max_row
    makers = []
    for i in range(2, workcell + 1):
        maker = sheet.cell(i, column).value
        if maker not in makers and maker is not None:
            makers.append(str(sheet.cell(i, column).value))

    wb.close()
    return makers


def get_products(maker, path="databases/catalog.xlsx"):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    workcell = sheet.max_row
    products = []

    for i in range(2, workcell + 1):
        if sheet.cell(i, 1).value == maker:
            if int(sheet.cell(i, 5).value) > 0:
                product = str(sheet.cell(i, 2).value)
                if product not in products:
                    products.append(product)

    return products


def find_product(maker, taste, path='databases/catalog.xlsx'):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    info = {}
    for i in range(1, sheet.max_row + 1):
        if str(maker).lower() in str(sheet.cell(i, 1).value).lower() and str(taste).lower() in str(
                sheet.cell(i, 2).value).lower():
            info['maker'] = sheet.cell(i,1).value
            info['taste'] = sheet.cell(i, 2).value
            info['puffs'] = sheet.cell(i, 3).value
            info['price'] = sheet.cell(i, 4).value
            info['count'] = sheet.cell(i, 5).value
            wb.close()
            return info
    return 'Not Found'


def catalog_get(path='databases/catalog.xlsx'):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    catalog = []
    for i in range(2, sheet.max_row + 1):
        print(str(type(sheet.cell(i,1).value)))
        if str(type(sheet.cell(i,1).value)) != "<class 'NoneType'>":
            product = {
                'maker': str(sheet.cell(i, 1).value).upper(),
                'taste': str(sheet.cell(i, 2).value),
                'puffs': int(sheet.cell(i, 3).value),
                'price': int(sheet.cell(i, 4).value),
                'count': int(sheet.cell(i, 5).value)
            }
            catalog.append(product)
    wb.close()
    return catalog


def catalog_change_count(maker, taste, num ,path ='databases/catalog.xlsx'):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    for i in range(1, sheet.max_row + 1):
        if str(maker).lower() in str(sheet.cell(i, 1).value).lower() and str(taste).lower() in str(
                sheet.cell(i, 2).value).lower():
            sheet.cell(i, 5).value = int(sheet.cell(i, 5).value) + num

            product = {
                'maker': str(sheet.cell(i, 1).value).upper(),
                'taste': str(sheet.cell(i, 2).value),
                'puffs': int(sheet.cell(i, 3).value),
                'price': int(sheet.cell(i, 4).value),
                'count': int(sheet.cell(i, 5).value)
            }
            wb.save(path)
            wb.close()
            return product

if __name__ == "__main__":
    print(find_product("HQD, Капучино"))
